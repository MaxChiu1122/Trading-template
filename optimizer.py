import pandas as pd
import numpy as np
from hyperopt import fmin, tpe, hp, Trials, STATUS_OK
from strategy import parse_strategy_logic, strategy_from_logic
from performance_metrics import calculate_performance_metrics
from excel_io import read_dashboard_inputs
from openpyxl import load_workbook


def write_optimization_results(excel_path, all_results):
    """Write optimization results (params and metrics) to the Optimization sheet."""
    wb = load_workbook(excel_path)
    # Write Optimization (test set) results as before
    ws_opt = wb["Optimization"]
    for row in ws_opt.iter_rows(min_row=1, max_row=ws_opt.max_row):
        for cell in row:
            cell.value = None
    param_keys = list(all_results[0]["BestParams"].keys())
    metric_keys = list(all_results[0]["TestMetrics"].keys())
    for key in ["PnL", "Equity Final [$]", "Equity Start [$]"]:
        if key in metric_keys:
            metric_keys.remove(key)
    headers = ["Window"] + param_keys + ["PnL"] + metric_keys
    for col_idx, col_name in enumerate(headers, start=1):
        ws_opt.cell(row=1, column=col_idx, value=col_name)
    for window_idx, result in enumerate(all_results):
        row = [f"Window {window_idx+1}"]
        row += [result["BestParams"].get(k, "") for k in param_keys]
        metrics = result["TestMetrics"].copy()
        pnl = metrics.get("PnL", None)
        if pnl is None or pnl == "":
            trades_df = result.get("TestTrades", None)
            if trades_df is not None and "PnL" in trades_df.columns:
                pnl = trades_df["PnL"].sum()
            else:
                eq_final = metrics.get("Equity Final [$]", None)
                eq_start = metrics.get("Equity Start [$]", None)
                if eq_final is not None and eq_start is not None:
                    pnl = eq_final - eq_start
                else:
                    pnl = ""
        row.append(pnl)
        for k in metric_keys:
            row.append(metrics.get(k, ""))
        for col_idx, val in enumerate(row, start=1):
            ws_opt.cell(row=window_idx+2, column=col_idx, value=val)

    # Write Train (train set) results if present
    if hasattr(write_optimization_results, "train_results") and write_optimization_results.train_results:
        train_results = write_optimization_results.train_results
        if "Train" not in wb.sheetnames:
            wb.create_sheet("Train")
        ws_train = wb["Train"]
        for row in ws_train.iter_rows(min_row=1, max_row=ws_train.max_row):
            for cell in row:
                cell.value = None
        # Use same headers as Optimization
        for col_idx, col_name in enumerate(headers, start=1):
            ws_train.cell(row=1, column=col_idx, value=col_name)
        for window_idx, result in enumerate(train_results):
            row = [f"Window {window_idx+1}"]
            row += [result["BestParams"].get(k, "") for k in param_keys]
            metrics = result["TrainMetrics"].copy()
            pnl = metrics.get("PnL", None)
            if pnl is None or pnl == "":
                trades_df = result.get("TrainTrades", None)
                if trades_df is not None and "PnL" in trades_df.columns:
                    pnl = trades_df["PnL"].sum()
                else:
                    eq_final = metrics.get("Equity Final [$]", None)
                    eq_start = metrics.get("Equity Start [$]", None)
                    if eq_final is not None and eq_start is not None:
                        pnl = eq_final - eq_start
                    else:
                        pnl = ""
            row.append(pnl)
            for k in metric_keys:
                row.append(metrics.get(k, ""))
            for col_idx, val in enumerate(row, start=1):
                ws_train.cell(row=window_idx+2, column=col_idx, value=val)
    wb.save(excel_path)


def metric_key_map():
    """Map Excel/weight keys to actual metric keys in metrics dict."""
    return {
        'AccReturn': 'Return [%]',
        'Sharpe': 'Sharpe Ratio',
        'Max Drawdown': 'Max Drawdown [%]',
        'Accuracy': 'Win Rate [%]',
        'SqrtMSE': 'SqrtMSE',
    }


def optimize_strategy(config):
    """Run rolling window optimization and return metrics, trades, and best params for each window."""
    excel_path = config.get("excel_path", "excel/trading_template.xlsx")
    df_all_orig = config["market_data"].copy()
    logic_df = config["logic_table"]
    param_ranges = config["param_ranges"]
    optimize_params = config["opt_params"]
    objective_weights = config.get("objective_weights", {})
    objective_type = config.get("objective_type", "MAX")
    train_window = int(config.get("train_window", 20))
    test_window = int(config.get("test_window", 5))
    max_evals = int(config.get("max_evals", 100))
    builder_df = config.get("indicator_builder")
    talib_df = config.get("talib_builder")
    print(f"🔍 optimize_params: {optimize_params}")
    print(f"🧪 param_ranges: {param_ranges}")
    print(f"🎯 objective_weights: {objective_weights}")
    if not param_ranges:
        print("❌ No valid parameter ranges found for optimization. Skipping optimization.")
        return pd.DataFrame(), [], []

    # Use shared build_indicators function
    from indicator_builder import build_indicators

    all_results = []
    best_params_list = []
    indicators_per_trial = []  # Store indicator DataFrame for each trial
    test_indicator_dfs = []  # Store full test set DataFrame (with indicators) for each window
    train_results = []  # Store train set metrics/trades for each window
    # Determine parameter types (float or int) based on step size
    param_types = {}
    for p in optimize_params:
        if p in param_ranges:
            step = param_ranges[p][2]
            if float(step).is_integer():
                param_types[p] = int
            else:
                param_types[p] = float

    for window_idx, start_idx in enumerate(range(0, len(df_all_orig) - train_window - test_window + 1, test_window)):
        # Always start from the original data for each window
        train_df = df_all_orig.iloc[start_idx:start_idx + train_window].copy()
        test_df = df_all_orig.iloc[start_idx + train_window:start_idx + train_window + test_window].copy()

        trial_indicator_snapshots = []  # Store indicator DataFrame for each trial in this window

        def objective(params):
            param_dict = {p: param_types[p](params[p]) for p in optimize_params}
            # Build indicators for this parameter set
            train_df_local = train_df.copy()
            train_df_local = build_indicators(train_df_local, param_dict, builder_df=builder_df, talib_df=talib_df)
            # Store a snapshot of indicators for this trial
            trial_indicator_snapshots.append((param_dict.copy(), train_df_local.copy()))
            rule_dict = parse_strategy_logic(logic_df)
            result_df = strategy_from_logic(train_df_local, rule_dict)
            metrics = calculate_performance_metrics(result_df, train_df_local)
            key_map = metric_key_map()
            score = 0
            for metric, weight in objective_weights.items():
                mapped_key = key_map.get(metric, metric)
                val = metrics.get(mapped_key, 0)
                score += weight * val
            if objective_type == "MAX":
                return {"loss": -score, "status": STATUS_OK}
            else:
                return {"loss": score, "status": STATUS_OK}

        search_space = {
            p: hp.quniform(p, *param_ranges[p])
            for p in optimize_params if p in param_ranges
        }
        trials = Trials()
        best = fmin(fn=objective, space=search_space, algo=tpe.suggest, max_evals=max_evals, trials=trials)
        # Cast best params to correct type
        best = {k: param_types[k](v) for k, v in best.items()}

        # --- Train set metrics/trades ---
        train_df_local = train_df.copy()
        train_df_local = build_indicators(train_df_local, best, builder_df=builder_df, talib_df=talib_df)
        rule_dict = parse_strategy_logic(logic_df)
        train_result_df = strategy_from_logic(train_df_local, rule_dict)
        train_metrics = calculate_performance_metrics(train_result_df, train_df_local)
        eq_final_tr = train_metrics.get("Equity Final [$]", None)
        eq_start_tr = train_metrics.get("Equity Start [$]", None)
        if eq_final_tr is not None and eq_start_tr is not None:
            pnl_tr = eq_final_tr - eq_start_tr
        else:
            pnl_tr = None
        train_metrics["PnL"] = pnl_tr
        if "Equity Final [$]" in train_metrics:
            del train_metrics["Equity Final [$]"]
        if "Equity Start [$]" in train_metrics:
            del train_metrics["Equity Start [$]"]
        train_results.append({
            "BestParams": best,
            "TrainMetrics": train_metrics,
            "TrainTrades": train_result_df
        })

        # --- Test set metrics/trades ---
        test_df_local = test_df.copy()
        test_df_local = build_indicators(test_df_local, best, builder_df=builder_df, talib_df=talib_df)
        # Save a copy of the test set DataFrame with all indicators
        test_indicator_dfs.append(test_df_local.copy())
        rule_dict = parse_strategy_logic(logic_df)
        test_result_df = strategy_from_logic(test_df_local, rule_dict)
        test_metrics = calculate_performance_metrics(test_result_df, test_df_local)
        eq_final = test_metrics.get("Equity Final [$]", None)
        eq_start = test_metrics.get("Equity Start [$]", None)
        if eq_final is not None and eq_start is not None:
            pnl = eq_final - eq_start
        else:
            pnl = None
        test_metrics["PnL"] = pnl
        if "Equity Final [$]" in test_metrics:
            del test_metrics["Equity Final [$]"]
        if "Equity Start [$]" in test_metrics:
            del test_metrics["Equity Start [$]"]
        all_results.append({
            "BestParams": best,
            "TestMetrics": test_metrics,
            "TestTrades": test_result_df
        })
        best_params_list.append(best)
        indicators_per_trial.append(trial_indicator_snapshots)
    # Attach train_results to the writer for later use
    write_optimization_results.train_results = train_results



    write_optimization_results(excel_path, all_results)
    key_map = metric_key_map()
    df_metrics = pd.DataFrame([r["TestMetrics"] for r in all_results])
    # Ensure all mapped keys exist in df_metrics
    for excel_key, metric_key in key_map.items():
        if metric_key in df_metrics.columns:
            df_metrics[excel_key] = df_metrics[metric_key]
        elif excel_key in df_metrics.columns:
            continue
        else:
            df_metrics[excel_key] = None
    # Also return all test trades for each window, best params, and full test indicator DataFrames
    test_trades_list = [r["TestTrades"] for r in all_results]
    return df_metrics, test_trades_list, best_params_list, test_indicator_dfs


# if __name__ == "__main__":
#     config = read_dashboard_inputs("excel/trading_template.xlsx")
#     config["excel_path"] = "excel/trading_template.xlsx"
#     results_df, test_trades_list, best_params_list, test_indicator_dfs = optimize_strategy(config)
#     print(results_df)
#     # Example: robust best parameter selection for a metric (e.g., 'AccReturn')
#     best_metric = 'AccReturn'  # or any metric you want to maximize
#     if best_metric in results_df.columns:
#         best_idx_label = results_df[best_metric].idxmax()
#         best_idx = results_df.index.get_loc(best_idx_label)
#         best_params = best_params_list[best_idx]
#         print(f"Best params for {best_metric}: {best_params}")
#     else:
#         print(f"Metric {best_metric} not found in results_df columns: {results_df.columns.tolist()}")


