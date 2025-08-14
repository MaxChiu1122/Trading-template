# --- Utility: Write best parameters back to Dashboard sheet ---
def write_best_params_to_dashboard(file_path: str, best_params: dict):
    """
    Update the Dashboard parameter table with the best parameters found by optimization.
    """
    wb = load_workbook(filename=file_path)
    ws = wb["Dashboard"]
    # Find the parameter table (look for 'Initial' anchor)
    value_row, value_col = find_anchor(ws, "Initial")
    r = value_row + 1 if value_row else 22
    updated = set()
    while True:
        key_cell = ws.cell(row=r, column=1).value
        if key_cell is None:
            break
        if "(" in str(key_cell) and ")" in str(key_cell):
            key = str(key_cell).split("(")[-1].strip(") ")
            if key in best_params:
                ws.cell(row=r, column=2, value=best_params[key])
                updated.add(key)
        r += 1
    wb.save(file_path)
    print(f"[INFO] Updated Dashboard parameters: {sorted(updated)}")
import pandas as pd
from openpyxl import load_workbook



def find_anchor(ws, anchor_text):
    """Find the cell (row, col) of the first cell containing anchor_text."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == anchor_text:
                return cell.row, cell.column
    return None, None

def extract_table(ws, anchor_text, max_cols=20, max_rows=5000):
    """Extract a table from ws starting at the anchor cell (header), until blank row."""
    start_row, start_col = find_anchor(ws, anchor_text)
    if start_row is None:
        raise ValueError(f"找不到 anchor '{anchor_text}'")
    # Read header
    headers = []
    for col in range(start_col, start_col + max_cols):
        val = ws.cell(row=start_row, column=col).value
        if val is None:
            break
        headers.append(str(val))
    # Read data rows
    data = []
    for r in range(start_row + 1, start_row + max_rows):
        row_vals = []
        empty_row = True
        for c in range(start_col, start_col + len(headers)):
            v = ws.cell(row=r, column=c).value
            row_vals.append(v)
            if v is not None:
                empty_row = False
        if empty_row:
            break
        data.append(row_vals)
    df = pd.DataFrame(data, columns=headers)
    return df

def read_dashboard_inputs(file_path: str) -> dict:
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb["Dashboard"]

    config = {}

    # --- Step 1: Extract market data table (anchor: "Date") ---
    market_df = extract_table(ws, "Date", max_cols=100, max_rows=5000)

    # --- Step 2: Extract parameter map (anchor: "VaInitiallue") ---
    value_row, value_col = find_anchor(ws, "Initial")
    param_map = {}
    param_ranges = {}
    r = value_row + 1 if value_row else 22
    while True:
        key_cell = ws.cell(row=r, column=1).value
        val_cell = ws.cell(row=r, column=2).value
        if key_cell is None:
            break
        if "(" in str(key_cell) and ")" in str(key_cell):
            key = str(key_cell).split("(")[-1].strip(") ")
            param_map[key] = val_cell
            # Range columns: C, D, E
            low = ws.cell(row=r, column=3).value
            high = ws.cell(row=r, column=4).value
            steps = ws.cell(row=r, column=5).value
            # Only add to param_ranges if all are valid numbers
            try:
                if low is not None and high is not None and steps is not None:
                    low_f = float(low)
                    high_f = float(high)
                    # steps: int if int, float if float
                    steps_val = float(steps)
                    if steps_val.is_integer():
                        steps_val = int(steps_val)
                    param_ranges[key] = (low_f, high_f, steps_val)
            except Exception:
                pass
        r += 1

    # --- Step 3: Extract Indicator Builder (anchor: "Indicator Name") ---
    builder_df = extract_table(ws, "Indicator Name", max_cols=5, max_rows=100)
    # print("[DEBUG] Arithmetic builder table:\n", builder_df)

    # --- Step 4: Extract Indicator Builder (TALib) (anchor: "TA-Lib Name") ---
    talib_df = extract_table(ws, "TA-Lib Name", max_cols=5, max_rows=100)
    # print("[DEBUG] TA-Lib builder table:\n", talib_df)


    # === Build indicators before logic extraction ===
    # Build arithmetic indicators from the builder table with combination logic
    if 'Combination' in builder_df.columns:
        # Group by Indicator Name, process each chain
        for name, group in builder_df.groupby('Indicator Name', sort=False):
            result = None
            for idx, row in group.iterrows():
                ind_a = row.get("Indicator A")
                op = row.get("Operator")
                val_or_param = row.get("Value / Param")
                comb = str(row.get("Combination", "END")).strip().upper()
                if not (name and ind_a and op and val_or_param):
                    continue
                if str(ind_a) not in market_df.columns:
                    continue
                val = param_map.get(str(val_or_param), val_or_param)
                if str(val) in market_df.columns:
                    val_operand = market_df[str(val)]
                else:
                    try:
                        val_operand = float(val)
                    except Exception:
                        print(f"[DEBUG] Skipping indicator '{name}': value '{val}' is not a valid number or column. Available columns: {list(market_df.columns)}")
                        continue
                # Compute this step
                try:
                    step_result = None
                    if op == "+":
                        step_result = market_df[str(ind_a)] + val_operand
                    elif op == "-":
                        step_result = market_df[str(ind_a)] - val_operand
                    elif op == "*":
                        step_result = market_df[str(ind_a)] * val_operand
                    elif op == "/":
                        step_result = market_df[str(ind_a)] / val_operand
                    elif op == "**":
                        step_result = market_df[str(ind_a)] ** val_operand
                    else:
                        continue
                    if result is None:
                        result = step_result
                    else:
                        # Combine with previous result
                        if comb == "+":
                            result = result + step_result
                        elif comb == "-":
                            result = result - step_result
                        elif comb == "*":
                            result = result * step_result
                        elif comb == "/":
                            result = result / step_result
                        elif comb == "END":
                            pass  # END, do nothing
                except Exception:
                    continue
            if result is not None:
                market_df[name] = result
    else:
        # Fallback: original logic for legacy sheets
        for _, row in builder_df.iterrows():
            name = row.get("Indicator Name")
            ind_a = row.get("Indicator A")
            op = row.get("Operator")
            val_or_param = row.get("Value / Param")
            # Skip if any required field is missing
            if not (name and ind_a and op and val_or_param):
                continue
            # Ensure the base indicator exists
            if str(ind_a) not in market_df.columns:
                continue
            # Determine the operand: either a column or a value/parameter
            val = param_map.get(str(val_or_param), val_or_param)
            if str(val) in market_df.columns:
                val_operand = market_df[str(val)]
            else:
                try:
                    val_operand = float(val)
                except Exception:
                    print(f"[DEBUG] Skipping indicator '{name}': value '{val}' is not a valid number or column. Available columns: {list(market_df.columns)}")
                    continue
            # Perform the arithmetic operation
            try:
                if op == "+":
                    market_df[name] = market_df[str(ind_a)] + val_operand
                elif op == "-":
                    market_df[name] = market_df[str(ind_a)] - val_operand
                elif op == "*":
                    market_df[name] = market_df[str(ind_a)] * val_operand
                elif op == "/":
                    market_df[name] = market_df[str(ind_a)] / val_operand
                elif op == "**":
                    market_df[name] = market_df[str(ind_a)] ** val_operand
            except Exception:
                continue

    # Build TA-Lib indicators after arithmetic indicators
    import talib
    for idx, row in talib_df.iterrows():
        name = row.get("TA-Lib Name")
        func = row.get("TA-Lib Function")
        in_col = row.get("In order Indicators")
        param_str = row.get("In order Param")
        if not (name and func):
            continue
        # Support comma-separated input columns for multi-input functions
        input_cols = [c.strip() for c in str(in_col).split(",") if c.strip()] if in_col else []
        input_series = []
        if input_cols and all(c in market_df.columns for c in input_cols):
            input_series = [market_df[c] for c in input_cols]
        else:
            continue
        param_keys = [p.strip() for p in str(param_str).split(",") if p.strip()] if param_str else []
        param_vals = [param_map.get(k, k) for k in param_keys]
        for i, v in enumerate(param_vals):
            try:
                param_vals[i] = int(float(v))
            except Exception:
                try:
                    param_vals[i] = float(v)
                except Exception:
                    pass
        try:
            talib_func = getattr(talib, func)
        except Exception:
            continue
        try:
            if len(input_series) == 1:
                out = talib_func(input_series[0], *param_vals)
            else:
                out = talib_func(*input_series, *param_vals)
            out_names = [n.strip() for n in name.split(",")]
            assigned = False
            if isinstance(out, (tuple, list)) and len(out_names) == len(out):
                if all(hasattr(o, '__len__') and not isinstance(o, str) and len(o) == len(market_df) for o in out):
                    for n, o in zip(out_names, out):
                        market_df[n] = o
                    assigned = True
            if not assigned:
                if hasattr(out, '__len__') and not isinstance(out, str) and len(out) == len(market_df):
                    market_df[name] = out
        except Exception as e:
            continue

    # Paste new indicators to the right of existing indicators in the Dashboard sheet
    # Assumes indicators start at column 'L' (12) and row 2
    ws_dashboard = ws
    indicator_start_col = 12  # 'L'
    indicator_row = 2
    # Find the last non-empty column in the indicator row
    last_col = indicator_start_col
    while ws_dashboard.cell(row=indicator_row, column=last_col).value is not None:
        last_col += 1
    # Get the list of original indicator columns
    original_cols = [ws_dashboard.cell(row=indicator_row, column=col).value for col in range(indicator_start_col, last_col)]
    # Identify new indicators to add (exclude standard OHLCV columns)
    exclude_cols = ["Date", "Open", "High", "Low", "Close", "Volume"]
    new_indicators = [col for col in market_df.columns if col not in original_cols and col not in exclude_cols]
    # Write new indicator headers and values
    for idx, ind in enumerate(new_indicators):
        ws_dashboard.cell(row=indicator_row, column=last_col + idx, value=ind)
        for row_idx, val in enumerate(market_df[ind], start=indicator_row + 1):
            ws_dashboard.cell(row=row_idx, column=last_col + idx, value=val)

    config["market_data"] = market_df
    config["param_map"] = param_map
    config["param_ranges"] = param_ranges
    config["indicator_builder"] = builder_df
    config["talib_builder"] = talib_df

    # --- Step 5: Extract Backtest Setting (anchor: "Duration") ---
    bt_row, bt_col = find_anchor(ws, "Duration")
    config["start_date"] = pd.to_datetime(ws.cell(row=bt_row + 1, column=bt_col + 1).value)
    config["end_date"] = pd.to_datetime(ws.cell(row=bt_row + 2, column=bt_col + 1).value)

    # --- Step 6: Extract Train-Test Settings (anchor: "Settings") ---
    train_row, train_col = find_anchor(ws, "Settings")
    config["train_window"] = int(ws.cell(row=train_row + 1, column=train_col + 1).value)
    config["test_window"] = int(ws.cell(row=train_row + 2, column=train_col + 1).value)

    # --- Step 7: Optimize Objection List (anchor: "Optimize Parameters") ---
    opt_row, opt_col = find_anchor(ws, "Optimize Parameters")
    opt_params_raw = ws.cell(row=opt_row, column=opt_col + 1).value
    if opt_params_raw:
        config["opt_params"] = [p.strip() for p in str(opt_params_raw).split(",") if p.strip()]
    else:
        config["opt_params"] = []
    config["objective_type"] = str(ws.cell(row=opt_row + 1, column=opt_col + 1).value).strip().upper() if ws.cell(row=opt_row + 1, column=opt_col + 1).value else "MAX"
    config["max_evals"] = int(ws.cell(row=opt_row + 2, column=opt_col + 1).value) if ws.cell(row=opt_row + 2, column=opt_col + 1).value else 100

    # --- Step 8: Objective metrics and weights (anchor: "Optimization Metric") ---
    obj_row, obj_col = find_anchor(ws, "Optimization Metric")
    objective_weights = {}
    for r in range(obj_row + 1, obj_row + 6):
        metric = ws.cell(row=r, column=obj_col).value
        weight = ws.cell(row=r, column=obj_col + 1).value
        if metric and weight is not None:
            objective_weights[str(metric).strip()] = float(weight)
    config["objective_weights"] = objective_weights

    # --- Step 9: Extract Strategy Logic Builder (anchor: "Rule Type") ---
    # Only extract and validate after market_data is updated with all indicators
    logic_df = extract_table(ws, "Rule Type", max_cols=7, max_rows=100)
    config["logic_table"] = logic_df

    # --- Step 10: Validate columns in logic table ---
    # Only validate after update_excel_with_market_data has run and all indicators are present
    # This function should only build and return config, not raise errors for missing columns
    valid_columns = set(config["market_data"].columns)
    invalid_rows = []
    for i, row in logic_df.iterrows():
        col_a = str(row["Column A"]).strip()
        col_b = str(row["Column B / Value"]).strip()
        if col_a not in valid_columns:
            invalid_rows.append((i + 2, f"Column A '{col_a}' 不存在於 market_data 中"))
    config["logic_table_invalid_rows"] = invalid_rows

    return config



def write_results(file_path: str, results_df: pd.DataFrame, metrics: dict = None):
    """
    Write backtest or optimization results and metrics to the Results sheet in the Excel file.
    Clears old results (except headers) and writes new results.
    """
    wb = load_workbook(filename=file_path)
    ws = wb["Results"]

    # Clear old result content (keep headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write new results DataFrame
    if not results_df.empty:
        for col_idx, col_name in enumerate(results_df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in results_df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)

    # Optionally write metrics to the right of the results table
    if metrics:
        # Find the first empty column after the results table
        start_col = results_df.shape[1] + 2  # leave one column gap
        ws.cell(row=1, column=start_col, value="Performance Metrics")
        for i, (k, v) in enumerate(metrics.items()):
            ws.cell(row=i + 2, column=start_col, value=k)
            ws.cell(row=i + 2, column=start_col + 1, value=v)

    wb.save(file_path)


def write_data_table(file_path: str, df: pd.DataFrame, sheet_name: str = "Data"):
    """
    Write a DataFrame to a specified sheet in the Excel file, replacing its content.
    """
    wb = load_workbook(filename=file_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    # Clear old content
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    # Write DataFrame
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx + 2, column=col_idx, value=value)
    wb.save(file_path)


# if __name__ == "__main__":
#     config = read_dashboard_inputs("excel/trading_template.xlsx")
#     print(config["logic_table"])
#     print(config["market_data"])