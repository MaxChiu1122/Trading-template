"""
Excel Trading Optimizer - Main Entry Point

This module provides the main workflow orchestration for the trading strategy
backtesting and optimization system. It coordinates between Excel configuration,
strategy execution, performance analysis, and visualization generation.

Key Features:
- Excel-driven configuration management
- Strategy backtesting with multi-period position support
- Rolling window parameter optimization
- Comprehensive performance analysis
- Automated visualization generation

Usage:
    python main.py                    # Run basic backtest
    main(optimize=True)              # Run parameter optimization
    main(optimize=False)             # Run single backtest

Author: Excel Trading Optimizer Project
License: MIT
"""

from excel_io import read_dashboard_inputs, write_results
from performance_metrics import calculate_performance_metrics
from generate_visuals import plot_visualization
from strategy import parse_strategy_logic, strategy_from_logic
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from optimizer import optimize_strategy
import pandas as pd


def insert_plot_into_excel(excel_path: str, image_path: str, sheet_name: str = "Visualization"):
    """Insert a plot image into the specified Excel sheet, replacing any existing images."""
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    # Remove existing images
    ws._images.clear()
    img = ExcelImage(image_path)
    img.anchor = 'A1'
    ws.add_image(img)
    wb.save(excel_path)


def main(optimize: bool = False):
    """Main entry point for running backtest or optimization workflow."""
    excel_path = "excel/trading_template.xlsx"
    # symbol = "ES=F"

    # Update market data and build indicators
    # update_excel_with_market_data(excel_path, symbol, download_data=False)

    # Read config and logic after market_data is updated
    config = read_dashboard_inputs(excel_path)
    config["excel_path"] = excel_path

    df = config["market_data"]
    print("[DEBUG] Columns in market_data after config:", df.columns.tolist())
    if optimize:
        print("Running optimization mode...")
        results_df, test_trades_list, best_params_list, test_indicator_dfs = optimize_strategy(config)
        all_trades = pd.concat(test_trades_list, ignore_index=True)
        combined_metrics = calculate_performance_metrics(all_trades, config["market_data"])
        write_results(excel_path, all_trades, combined_metrics)

        # Find best parameter set by objective
        best_metric = next(iter(config["objective_weights"].keys()))
        # Robust handling for all-NA or missing best_metric
        if best_metric not in results_df.columns:
            print(f"[ERROR] Metric '{best_metric}' not found in results_df columns: {results_df.columns.tolist()}")
            best_params = best_params_list[0] if best_params_list else {}
        elif results_df[best_metric].isna().all():
            print(f"[ERROR] All values for metric '{best_metric}' are NA. Cannot select best parameters.")
            best_params = best_params_list[0] if best_params_list else {}
        else:
            if config["objective_type"] == "MAX":
                best_idx_label = results_df[best_metric].idxmax()
            else:
                best_idx_label = results_df[best_metric].idxmin()
            if pd.isna(best_idx_label):
                print(f"[ERROR] idxmax/idxmin returned NaN for metric '{best_metric}'. Using first parameter set as fallback.")
                best_params = best_params_list[0] if best_params_list else {}
            else:
                best_idx = results_df.index.get_loc(best_idx_label)
                best_params = best_params_list[best_idx]
        config.update(best_params)

        # --- Completely rewrite Data sheet: only test set rows, include all indicators and params ---
        param_cols = list(best_params_list[0].keys())
        test_rows = []
        for i, test_df in enumerate(test_indicator_dfs):
            if test_df is not None and not test_df.empty:
                # Add *_used columns for this window
                for k in param_cols:
                    test_df[f"{k}_used"] = best_params_list[i][k]
                test_rows.append(test_df)
        if test_rows:
            df_data = pd.concat(test_rows, ignore_index=True)
        else:
            df_data = pd.DataFrame()

        # Ensure all base columns are present by merging with original market data
        base_cols = ["Date", "Open", "High", "Low", "Close", "Volume", "Pt"]
        market_base = config["market_data"][base_cols].copy() if all(col in config["market_data"].columns for col in base_cols) else config["market_data"].copy()
        # Merge on Date, giving priority to test set values
        if not df_data.empty:
            df_data = pd.merge(df_data, market_base, on="Date", how="left", suffixes=("", "_mkt"))
            # For each base col, if missing in df_data, fill from market_base
            for col in base_cols:
                if col not in df_data.columns:
                    df_data[col] = df_data[f"{col}_mkt"]
            # Remove any *_mkt columns
            df_data = df_data[[c for c in df_data.columns if not c.endswith("_mkt")]]
        else:
            df_data = market_base.iloc[0:0].copy()

        # Reorder columns: Date, Open, High, Low, Close, Volume, Pt, indicators, *_used
        base_cols_present = [col for col in base_cols if col in df_data.columns]
        indicator_cols = [col for col in df_data.columns if col not in base_cols_present and not col.endswith("_used")]
        used_cols = [col for col in df_data.columns if col.endswith("_used")]
        ordered_cols = base_cols_present + indicator_cols + used_cols
        df_data = df_data[ordered_cols]

        # Visualization (use the test set rows)
        png_path = plot_visualization(df_data, all_trades, output_folder="images")
        insert_plot_into_excel(excel_path, png_path, sheet_name="Visualization")

        from excel_io import write_data_table
        # Write only the test set rows to the Data sheet
        write_data_table(excel_path, df_data, sheet_name="Data")
    else:
        # Run normal backtest
        logic_df = config["logic_table"]
        rule_dict = parse_strategy_logic(logic_df)
        result_df = strategy_from_logic(df, rule_dict)
        metrics = calculate_performance_metrics(result_df, df)
        write_results(excel_path, result_df, metrics)
        png_path = plot_visualization(df, result_df, output_folder="images")
        insert_plot_into_excel(excel_path, png_path, sheet_name="Visualization")

    # # Optional: open interactive HTML
    # if Path(html_path).exists():
    #     webbrowser.open(f"file://{Path(html_path).resolve()}")


if __name__ == "__main__":
    # Set optimize=True to run optimization, False for normal backtest
    main(optimize=False)